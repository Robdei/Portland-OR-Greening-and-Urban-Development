from PIL import Image
import glob
import openpyxl
import math

#data=pd.read_excel('Portland Greening.xlsx',sheetname='Landsat Data')
#ML=data['ML10'].tolist()
files=[]
for i in glob.glob('*.tif'):
	files.append(i)

Positions=[[3071,4853+2*489],[3020,4851+2*489],[3061,4851+2*489],[3135,4859+2*489],[3068,4852+2*489],[3022,4852+2*489],[3053,4851+2*489],[3076,4855+2*489],[3090,4864+2*489],[3097,4855+2*489]]

ML=0.00002
AL=-.1

def radiance(Intensity):
	return(ML*Intensity+AL)

#print(files)

wb=openpyxl.load_workbook('TempVDistance.xlsx')
#for file in files:
#	wb.create_sheet(file)
#wb.save('TempVDistance.xlsx')
col=1
for i in range(len(files)):
	col=col+1
	im=Image.open(files[i])

	pix=im.load()
	ws=wb['Green Reflect']


	w,h=im.size
	Pioneer=Positions[i]

	count=-1
	for x in range(Positions[i][0],w):
		count=count+1
		ws.cell(row=x+2-Positions[i][0],column=col).value=radiance(pix[x,Positions[i][1]])
		ws.cell(row=x+2-Positions[i][0],column=1).value=count*.03

	wb.save('TempVDistance.xlsx')
