from PIL import Image
import glob, os
import openpyxl
from scipy.interpolate import CubicSpline
import numpy as np
import matplotlib as plt
import math

wb=openpyxl.load_workbook('Portland Greening.xlsx')
ws2=wb['Landsat Data']
ws=wb['Colorspace Conversion']

ML=[]
for i in range(2,14):
	ML.append(ws.cell(row=i,column=4).value)

AL=[]
for i in range(17,29):
	ML.append(ws.cell(row=i,column=4).value)
print(ML)
print(AL)
def radiance(Intensity,index):
	return(ML[index]*Intensity+AL[index])

def Temp(Radiance):
	return(k2/(math.log(k1/Radiance+1)))


Tempp,Redpix,Greenpix,Bluepix=[np.zeros(54),np.zeros(54),np.zeros(54),np.zeros(54)]
#print(Tempp)
for i in range(54):
	Tempp[i]=ws.cell(row=i+2,column=11).value
	Redpix[i]=ws.cell(row=i+2,column=12).value
	Greenpix[i]=ws.cell(row=i+2,column=13).value
	Bluepix[i]=ws.cell(row=i+2,column=14).value

#print(Tempp)
#Tempp=np.asarray(Tempp)
#Redpix=np.asarray(Redpix)
#Bluepix=np.asarray(Bluepix)
#Greenpix=np.asarray(Greenpix)
Red = CubicSpline(Tempp, Redpix)
Blue = CubicSpline(Tempp, Bluepix)
Green = CubicSpline(Tempp, Greenpix)

os.chdir('C:/Users/administrator/Desktop/Portland Greening/Images/Band3 (Green)')
files=[]
for i in glob.glob('*.tif'):
	files.append(i)
#print(files)
j=-1
for file in files:
	im=Image.open(file)
	pix=im.load()
	w,h=im.size
	j=j+1
	low=radiance(pix[w/2,h/2],j)
	high=radiance(pix[w/2,h/2],j)
	#im2=Image.new('RGB',im.size)
	#pix2=im2.load()

	for x in range(w):
		for y in range(h):
			if pix[x,y]==0:
				continue
			else:
				Pixtemp=radiance(pix[x,y],j)
				if Pixtemp>high:
					high=Pixtemp
				elif Pixtemp<low:
					low=Pixtemp
				#R=int(Red(Pixtemp))
				#G=int(Green(Pixtemp))
				#B=int(Blue(Pixtemp))
					
				#print([R,G,B])
				#pix2[x,y]=(R,G,B)
				print([file,low,high])
	print('In progress....')
	im2.save(file[:-4]+'Temp.png')