from PIL import Image
import glob
import openpyxl
import math

#data=pd.read_excel('Portland Greening.xlsx',sheetname='Landsat Data')
#ML=data['ML10'].tolist()
files=[]
for i in glob.glob('*.tif'):
	files.append(i)

Positions=[[2862+218,5339+489],[2843+218,5340+489],[2852+218,5341+489],[3032+218,5327+489],[2823+218,5342+489],[2853+218,5341+489],[3020,4853+2*489],[3071,4853+2*489],[3137,4859+2*489],[3020,4851+2*489],[3061,4851+2*489],[3135,4859+2*489],[3068,4852+2*489],[3022,4852+2*489],[3053,4851+2*489],[3076,4855+2*489],[3090,4864+2*489],[3097,4855+2*489]]

ML=0.0003342
AL=0.1
k1=774.8853
k2=1321.0789
def radiance(Intensity):
	return(ML*Intensity+AL)
def Temp(Radiance):
	return(k2/(math.log(k1/Radiance+1)))
#print(files)

Direction=['E','NE','N','NW','W','SW','S','SE']

wb=openpyxl.load_workbook('TempVDistance 2.xlsx')
for file in files:
	for d in Direction:
		wb.create_sheet(file[:-4]+d)
wb.save('TempVDistance 2.xlsx')

for i in range(len(files)):
	for d in Direction:
	
		im=Image.open(files[i])
		pix=im.load()
		ws=wb[files[i][:-4]+d]
		Pioneer=Positions[i]
		w,h=im.size
		if d=='E':
			count=-1
			for x in range(Positions[i][0],Positions[i][0]+2001):
				count=count+1
				ws.cell(row=x+2-Positions[i][0],column=2).value=Temp(radiance(pix[x,Positions[i][1]]))
				ws.cell(row=x+2-Positions[i][0],column=1).value=count*.03
				#pix[x,Positions[i][1]]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='NE':
			count=1
			for x in range(Positions[i][0],Positions[i][0]+2001):
				count=count-1
				ws.cell(row=x+2-Positions[i][0],column=2).value=Temp(radiance(pix[x,Positions[i][1]+count]))
				ws.cell(row=x+2-Positions[i][0],column=1).value=-count*(0.0424264)
				#pix[x,Positions[i][1]+count]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='N':
			count=-1
			for y in range(Positions[i][1],Positions[i][1]+2001):
				count=count+1
				ws.cell(row=y+2-Positions[i][1],column=2).value=Temp(radiance(pix[Positions[i][0],2*Positions[i][1]-y]))
				ws.cell(row=y+2-Positions[i][1],column=1).value=count*.03
				#pix[Positions[i][0],2*Positions[i][1]-y]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='S':
			count=-1
			for y in range(Positions[i][1],Positions[i][1]+2001):
				count=count+1
				ws.cell(row=y+2-Positions[i][1],column=2).value=Temp(radiance(pix[Positions[i][0],y]))
				ws.cell(row=y+2-Positions[i][1],column=1).value=count*.03
				#pix[Positions[i][0],y]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='NW':
			count=1
			for x in range(Positions[i][0],Positions[i][0]+2001):
				count=count-1
				ws.cell(row=x+2-Positions[i][0],column=2).value=Temp(radiance(pix[Positions[i][0]+count,Positions[i][1]+count]))
				ws.cell(row=x+2-Positions[i][0],column=1).value=-count*0.0424264
				#pix[Positions[i][0]+count,Positions[i][1]+count]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='W':
			count=1
			for x in range(Positions[i][0],Positions[i][0]+2001):
				count=count-1
				ws.cell(row=x+2-Positions[i][0],column=2).value=Temp(radiance(pix[Positions[i][0]+count,Positions[i][1]]))
				ws.cell(row=x+2-Positions[i][0],column=1).value=-count*.03
				#pix[Positions[i][0]+count,Positions[i][1]]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='SW':
			count=1
			for x in range(Positions[i][0],Positions[i][0]+2001):
				count=count-1
				ws.cell(row=x+2-Positions[i][0],column=2).value=Temp(radiance(pix[Positions[i][0]+count,Positions[i][1]-count]))
				ws.cell(row=x+2-Positions[i][0],column=1).value=-count*0.0424264
				#pix[Positions[i][0]+count,Positions[i][1]-count]=0
			#im.save(files[i][:-4]+d+'.tif')
		if d=='SE':
			count=1
			for x in range(Positions[i][0],Positions[i][0]+2001):
				count=count-1
				ws.cell(row=x+2-Positions[i][0],column=2).value=Temp(radiance(pix[Positions[i][0]-count,Positions[i][1]-count]))
				ws.cell(row=x+2-Positions[i][0],column=1).value=-count*0.0424264
				#pix[Positions[i][0]-count,Positions[i][1]-count]=0
			#im.save(files[i][:-4]+d+'.tif')

		wb.save('TempVDistance 2.xlsx')
		print('...........')