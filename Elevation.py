import numpy as np
import openpyxl
from scipy.interpolate import CubicSpline
import numpy as np

wb=openpyxl.load_workbook('Topography.xlsx')
ws=wb['Sheet1']
ws2=wb['NormalTopo']
D1=[]
E=[]
N=[]
W=[]
D2=[]
NE=[]
NW=[]
SW=[]
SE=[]
for i in range(2,263):
	D1.append(ws.cell(row=i,column=1).value)
	E.append(ws.cell(row=i,column=2).value)
	N.append(ws.cell(row=i,column=3).value)
	W.append(ws.cell(row=i,column=4).value)
for i in range(2,187):
	D2.append(ws.cell(row=i,column=6).value)
	NE.append(ws.cell(row=i,column=7).value)
	NW.append(ws.cell(row=i,column=8).value)
	SW.append(ws.cell(row=i,column=9).value)
	SE.append(ws.cell(row=i,column=10).value)
D1=np.array(D1)
E=np.array(E)
N=np.array(N)
W=np.array(W)
D2=np.array(D2)
NE=np.array(NE)
NW=np.array(NW)
SW=np.array(SW)
SE=np.array(SE)

EE = CubicSpline(D1, E)
NN = CubicSpline(D1, N)
WW = CubicSpline(D1, W)
NNE = CubicSpline(D2, NE)
NNW = CubicSpline(D2, NW)
SSW = CubicSpline(D2, SW)
SSE = CubicSpline(D2, SE)
count=1
for i in range(1,2004):
	count=count+1
	x=(count-2)*.03
	ws2.cell(row=count,column=1).value=x
	#print(EE(x))
	ws2.cell(row=count,column=2).value=int(EE(x))
	ws2.cell(row=count,column=3).value=int(NN(x))
	ws2.cell(row=count,column=4).value=int(WW(x))
	#ws2.cell(row=count,column=5).value=int(NNE(x))
	#ws2.cell(row=count,column=6).value=int(NNW(x))
	#ws2.cell(row=count,column=7).value=int(SSW(x))
	#ws2.cell(row=count,column=8).value=int(SSE(x))
count=1
for i in range(1,1422):
	count=count+1
	x=(count-2)*0.0424264
	ws2.cell(row=count,column=6).value=x
	#print(EE(x))
	#ws2.cell(row=count,column=11).value=int(EE(x))
	#ws2.cell(row=count,column=12).value=int(NN(x))
	#ws2.cell(row=count,column=13).value=int(WW(x))
	ws2.cell(row=count,column=7).value=int(NNE(x))
	ws2.cell(row=count,column=8).value=int(NNW(x))
	ws2.cell(row=count,column=9).value=int(SSW(x))
	ws2.cell(row=count,column=10).value=int(SSE(x))
wb.save('Topography.xlsx')




