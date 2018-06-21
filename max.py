from PIL import Image
import openpyxl

wb=openpyxl.load_workbook('Topography.xlsx')
ws=wb['Sheet1']

im=Image.open('topo.tif')

im.mode='I'
pix=im.load()
w,h=im.size
im2=im.convert('RGB')
pix2=im2.load()
count=1
for x in range(263):
	count=count+1
	ws.cell(row=count,column=2).value=pix[1108-567+(count-2),114+328]
	ws.cell(row=count,column=1).value=(count-2)*.23165
count=1
for y in range(263):
	count=count+1
	ws.cell(row=count,column=3).value=pix[1108-567,114+328-(count-2)]
count=1
for x in range(263):
	count=count+1
	ws.cell(row=count,column=4).value=pix[1108-567-(count-2),114+328]
	pix2[1108-567+(count-2),114+328]=(255,0,0)


count=1
for x in range(185):
	count=count+1
	ws.cell(row=count,column=10).value=pix[1108-567+(count-2),114+328+(count-2)]
	ws.cell(row=count,column=9).value=pix[1108-567-(count-2),114+328+(count-2)]
	ws.cell(row=count,column=8).value=pix[1108-567-(count-2),114+328-(count-2)]
	ws.cell(row=count,column=7).value=pix[1108-567+(count-2),114+328-(count-2)]
	ws.cell(row=count,column=6).value=(count-2)*.3276
wb.save('Topography.xlsx')
im2.save('tt.png')