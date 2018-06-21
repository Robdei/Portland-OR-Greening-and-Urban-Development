from PIL import Image
import glob, os
import openpyxl
from scipy.interpolate import CubicSpline
import numpy as np
import matplotlib as plt
import math



wb=openpyxl.load_workbook('Portland Greening.xlsx')
ws=wb['Landsat Data']

ML,AL,thetas=[[],[],{}]
Mp=.00002
Ap=-.10
ThermML=0.0003342
ThermAL=0.1
k1=774.8853
k2=1321.0789
folders=['20150810','20150927','20160406','20160812','20160913','20170103','20170527','20170714','20170730','20170815','20171205','20180311']

def GetBandStatistics(Date):
	for i in range(len(folders)):
		if folders[i]==Date:
			for j in range(2,13):
				ML.append(ws.cell(row=i+2,column=j).value)
			for i in range(2,13):
				AL.append(ws.cell(row=i+17,column=j).value)

for i in range(59,71):
	thetas[folders[i-59]]=(ws.cell(row=i,column=6).value)

def rho(Theta,Pixvalue):
	return((Mp*Pixvalue+Ap)/math.sin(Theta*np.pi/180))

def radiance(Ml,Al,Pixvalue):
	return(Ml*Pixvalue+Al)

def Temp(Radiance):
	return(k2/(math.log(k1/Radiance+1)))

#Band 4
def Filter1(Theta,Pixvalue):
	filtervalue=rho(Theta,Pixvalue)
	if filtervalue>.8:
		#Pass to filter 3
		return(True)
	else:
		#Pass to filter 2
		return(False)

#Band 4
def Filter2(Theta,Pixvalue):
	filtervalue=rho(Theta,Pixvalue)
	if filtervalue>=.7:
		return('ambiguous')
	else:
		#Not cloud
		return('Not Cloud')

#Band 3 & 6
def Filter3(Theta,GreenPixvalue,SWIRPixvalue):
	Green=rho(Theta,GreenPixvalue)
	SWIR=rho(Theta,SWIRPixvalue)
	NDSI=(Green-SWIR)/(Green+SWIR)
	if NDSI>-.25 and NDSI<.7:
		#Pass to filter 5
		return(True)
	else:
		#Pass to filter 4
		return(False)

#Band 3 & 6
def Filter4(Theta,GreenPixvalue,SWIRPixvalue):
	Green=rho(Theta,GreenPixvalue)
	SWIR=rho(Theta,SWIRPixvalue)
	NDSI=(Green-SWIR)/(Green+SWIR)
	if NDSI>.8:
		return('Ice')
	else:
		return('Dull Ice')

#Band 10
def Filter5(Pixvalue):
	if Temp(radiance(ThermML,ThermAL,Pixvalue))>300:
		#not cloud
		return('Not Cloud')
	else:
		#filter 6
		return(True)
#Band 6 and 10
def Filter6(Theta,SWIRPixvalue,ThermalPixvalue):
	tem=Temp(radiance(ThermML,ThermAL,ThermalPixvalue))
	ref=rho(Theta,SWIRPixvalue)
	if (1-ref)*tem<225:
		#Pass to filter 8
		return(True)
	else:
		#Pass to filter 7
		return(False)

#Band 6
def Filter7(Theta,SWIRPixvalue):
	ref=rho(Theta,SWIRPixvalue)
	if ref>=.8:
		return('ambiguous')
	else:
		#Not cloud
		return('Not cloud')

#Band 5 and 4
def Filter8(Theta,NIRPixvalue,RedPixvalue):
	redref=rho(Theta,RedPixvalue)
	NIRref=rho(Theta,NIRPixvalue)
	filtervalue=NIRref/redref
	if filtervalue>2.35:
		return(True)
	else:
		#Pass to filter 9
		return(False)

def Filter9(Theta,NIRPixvalue,GreenPixvalue):		
	greenref=rho(Theta,GreenPixvalue)
	NIRref=rho(Theta,NIRPixvalue)
	filtervalue=NIRref/greenref
	if filtervalue>2.16248:
		return(True)
	else:
		#Pass to filter 10
		return(False)

def Filter10(Theta,NIRPixvalue,SWIRPixvalue):
	SWIRref=rho(Theta,SWIRPixvalue)
	NIRref=rho(Theta,NIRPixvalue)
	if NIRref/SWIRref<1:
		return('ambiguous')
	else:
		return('Cloud')

def path(root):
	return('C:/Users/administrator/Desktop/Portland Greening/Images/'+root)

folders=['20150810','20150927','20160406','20160812','20160913','20170103','20170527','20170714','20170730','20170815','20171205','20180311']
Photonames=['_B3.tif','_B4.tif','_B5.tif','_B6.tif','B10.tif']

def Detector(Date,x,y):
	Filter_1=Filter1(thetas[Date],redpix[x,y])
	if Filter_1:
		Filter_3=Filter3(thetas[Date],greenpix[x,y],SWIRpix[x,y])
		if Filter_3:
			Filter_5=Filter5(LWIRpix[x,y])
			if Filter_5:
				Filter_6=Filter6(thetas[Date],SWIRpix[x,y],LWIRpix[x,y])
				if Filter_6:
					Filter_8=Filter8(thetas[Date],NIRpix[x,y],redpix[x,y])
					if not Filter_8:
						Filter_9=Filter9(thetas[Date],NIRpix[x,y],greenpix[x,y])
						if not Filter_9:
							return(Filter10(thetas[Date],NIRpix[x,y],SWIRpix[x,y]))
						if Filter_9:
							return('ambiguous')
					else:
						return('ambiguous')
				elif not Filter_6:
					return(Filter7(thetas[Date],SWIRpix[x,y]))
			else:
				return('Not Cloud')
		elif not Filter_3:
			return('Not Cloud')
	elif not Filter_1:
		return(Filter2(thetas[Date],redpix[x,y]))

def ACCA(Date):
	w,h=green.size
	#GetBandStatistics(Date)
	for x in range(w):
		for y in range(h):
			pixtype=Detector(Date,x,y)
			if pixtype=='Cloud':
				pixels[x,y]=(155,100,30)
			elif pixtype=='ambiguous':
				pixels[x,y]=(255,0,0)
			else:
				continue
	im.save('Cloudtest.tif')

os.chdir(path(folders[0]))
green=Image.open(folders[0]+Photonames[0])
im=Image.new('RGB',green.size,(255,255,255))
pixels=im.load()
red=Image.open(folders[0]+Photonames[1])
NIR=Image.open(folders[0]+Photonames[2])
SWIR=Image.open(folders[0]+Photonames[3])
LWIR=Image.open(folders[0]+Photonames[4])
greenpix=green.load()
redpix=red.load()
NIRpix=NIR.load()
SWIRpix=SWIR.load()
LWIRpix=LWIR.load()
ACCA(folders[0])


	

