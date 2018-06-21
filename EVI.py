from PIL import Image
import glob
import openpyxl
import math

def kmtopix(d):
	return(int(d/.03))

wb=openpyxl.load_workbook('Portland Greening.xlsx')
wb2=openpyxl.load_workbook('Green.xlsx')
#ML=data['ML10'].tolist()

files=[]
for i in glob.glob('*.tif'):
	files.append(i)
ws=wb['Landsat 5']
Positions=[]
for r in range(2,62):
	XPosition=ws.cell(row=r,column=16).value
	YPosition=ws.cell(row=r,column=17).value
	Positions.append([XPosition,YPosition])

for i in range(len(files)):
	print(files[i])
	im=Image.open(files[i])
	im.mode='I'
	pix=im.load()
	w,h=im.size
	Average=0
	ws2=wb2['Sheet1']
	#print(Positions[i])
	for x in range(Positions[i][0],Positions[i][0]+kmtopix(29.95)):
		for y in range(Positions[i][1]-kmtopix(9),Positions[i][1]+kmtopix(10)):
			Average=Average+pix[x,y]
			#if (x+y)%3000==0:
				#print(pix[x,y])
	for x in range(Positions[i][0]-kmtopix(17.5),Positions[i][0]+kmtopix(29.25)):
		for y in range(Positions[i][1]+kmtopix(10),Positions[i][1]+kmtopix(19.7)):
			Average=Average+pix[x,y]
	Average=Average/(kmtopix(19)*kmtopix(29.25)+kmtopix(9.7)*kmtopix(47.71))
	ws2.cell(row=i+1,column=1).value=files[i][:-4]
	ws2.cell(row=i+1,column=2).value=Average
wb2.save('Green.xlsx')
#im.save(files[i][:-4]+'test.png')




