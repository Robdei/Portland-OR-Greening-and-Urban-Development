import os, sys, glob
from PIL import Image
import openpyxl
from collections import Counter

wb=openpyxl.load_workbook('Green.xlsx')
ws=wb['Sheet1']
count=0
for file in glob.glob('*.tif'):
	count=count+1
	direc=[]
	im=Image.open(file)
	#for k in range(101):
	#	direc[k/100]=0
	im.mode='I'
	print(file)
	pix=im.load()
	w,h=im.size
	#kk=0
	for x in range(w):
		for y in range(h):
			#kk+1
			direc.append(round(pix[x,y],2)/10000)
			#if kk%10000000==0:
			#	print('..')

	#frequency=np.zeros(100)
	#maxx=0
	#for k in range(100):
	#	check=k/100
	#	for i in direc:
	#		if i==check:
	#			frequency[k]=frequency[k]+1
	#frequency.tolist()
	print(direc)
	data=Counter(direc)
	#for l in range(100):
	#	if frequency[l]>=maxx:
	#		maxx=l/100
	ws.cell(row=count,column=4).value=file[:-4]
	ws.cell(row=count,column=2).value=data.most_common(1)[0][0]
	ws.cell(row=count,column=3).value=data.most_common(2)[0][0]
wb.save('Green.xlsx')




			