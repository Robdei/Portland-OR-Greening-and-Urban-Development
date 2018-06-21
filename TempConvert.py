from PIL import Image
import glob, os
import openpyxl
from scipy.interpolate import CubicSpline
import numpy as np
import matplotlib as plt
import math

wb=openpyxl.load_workbook('Portland Greening.xlsx')
ws=wb['Colorspace Conversion']
lows=[224.39,264.41,248.69,274.8,267.1,240.29,271.2,274.53,274.94,231.43,255.32,259.9]
highs=[314.65,311.08,303.91,324.87,315.88,286.16,321.4,322.57,325.26,318.29,289.35,303.11]
Tempp=[]
Redpix=[]
Bluepix=[]
Greenpix=[]
ML=0.0003342
AL=0.1
k1=774.8853
k2=1321.0789
def radiance(Intensity):
	return(ML*Intensity+AL)

def Temp(Radiance):
	return(k2/(math.log(k1/Radiance+1)))

def Redpixel(Temp):
	if Temp<=6500 and Temp>=1000:
		return(255)
	elif Temp<1000:
		return(False)
	else:
		return(-7*10**-12*Temp**3+6*10**-7*Temp**2-.018*Temp+332.82)	
def Bluepixel(Temp):
	if Temp<2000 and Temp>=1000:
		return(0)
	elif Temp<1000:
		return(False)
	elif Temp>=2000 and Temp<6000:
		return(3*10**-9*Temp**3-5*10**-5*Temp**2+.2679*Temp-359.44)
	else:
		return(255)
def Greenpixel(Temp):
	if Temp<1000:
		return(False)
	elif Temp<6400 and Temp>1000:
		return(2*10**-9*Temp**3-2*10**-5*Temp**2+.1334*Temp-46.77)
	else:
		return(-4*10**-12*Temp**3+3*10**-7*Temp**2-.0101*Temp+292.05)
Tempp,Redpix,Greenpix,Bluepix=[np.zeros(54),np.zeros(54),np.zeros(54),np.zeros(54)]
#print(Tempp)
for i in range(54):
	Tempp[i]=ws.cell(row=i+2,column=11).value
	Redpix[i]=ws.cell(row=i+2,column=12).value
	Greenpix[i]=ws.cell(row=i+2,column=13).value
	Bluepix[i]=ws.cell(row=i+2,column=14).value

#print(Tempp)
#Tempp=np.asarray(Tempp)
#Redpix=np.asarray(Redpix)
#Bluepix=np.asarray(Bluepix)
#Greenpix=np.asarray(Greenpix)
Red = CubicSpline(Tempp, Redpix)
Blue = CubicSpline(Tempp, Bluepix)
Green = CubicSpline(Tempp, Greenpix)
print(int(Red(300)))
#print(Redpix)
os.chdir('C:/Users/administrator/Desktop/Portland Greening/Images/Band10')
files=[]
for i in glob.glob('*.tif'):
	files.append(i)
#print(files)
for file in files:
	im=Image.open(file)
	pix=im.load()
	w,h=im.size
	im2=Image.new('RGB',im.size)
	pix2=im2.load()

	for x in range(w):
		for y in range(h):
			if pix[x,y]==0:
				continue
			else:
				Pixtemp=Temp(radiance(pix[x,y]))
				R=int(Red(Pixtemp))
				G=int(Green(Pixtemp))
				B=int(Blue(Pixtemp))
				if R>255:
					R=255
				if G>255:
					G=255
				if B>255:
					B=255
				if R<0:
					R=0
				if G<0:
					G=0
				if B<0:
					B=0		
				#print([R,G,B])
				pix2[x,y]=(R,G,B)
	print('In progress....')
	im2.save(file[:-4]+'Temp.png')