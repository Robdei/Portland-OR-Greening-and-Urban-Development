from PIL import Image
import openpyxl
import numpy as np
#from scipy.interpolate import CubicSpline

Interval=int((326-220)/2)
im=Image.new('RGB',[Interval,45])
pix=im.load()

wb=openpyxl.load_workbook('Portland Greening.xlsx')
ws=wb['Colorspace Conversion']
Tempp,Redpix,Greenpix,Bluepix=[np.zeros(54),np.zeros(54),np.zeros(54),np.zeros(54)]

for i in range(54):
	Tempp[i]=ws.cell(row=i+2,column=11).value
	Redpix[i]=ws.cell(row=i+2,column=12).value
	Greenpix[i]=ws.cell(row=i+2,column=13).value
	Bluepix[i]=ws.cell(row=i+2,column=14).value

w,h=im.size
for x in range(w):
	for y in range(h):
		pix[x,y]=(int(Redpix[x]),int(Greenpix[x]),int(Bluepix[x]))
im.save('Legend.png')