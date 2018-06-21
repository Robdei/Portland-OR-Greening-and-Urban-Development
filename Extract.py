import openpyxl
import os, glob


wb=openpyxl.load_workbook('Portland Greening.xlsx')
ws=wb['Landsat 5']
direc=[]
for path, i, files in os.walk('C:/Users/administrator/Desktop/Portland Greening/Bulk Order 903560/Landsat 4-5 TM C1 Level-1/Ims'):
	direc.append(i)

#print(direc)
for k in range(len(direc[0])):
	os.chdir('C:/Users/administrator/Desktop/Portland Greening/Bulk Order 903560/Landsat 4-5 TM C1 Level-1/Ims'+'/'+direc[0][k])
	for i in glob.glob('*.txt'):
		if i[-7:]=='MTL.txt':
			#print(i)
			file = open(i, 'r') 
			text=file.read()
			for let in range(len(text)-22):
				a=text[let:let+22]
				b=text[let:let+23]
				if a in ['REFLECTANCE_ADD_BAND_1','REFLECTANCE_ADD_BAND_3','REFLECTANCE_ADD_BAND_4']:
					if a[-1]=='1':
						ws.cell(row=k+54,column=5).value=text[let+25:let+35]
					if a[-1]=='3':
						ws.cell(row=k+54,column=6).value=text[let+25:let+35]
					if a[-1]=='4':
						ws.cell(row=k+54,column=7).value=text[let+25:let+35]
				if b in ['REFLECTANCE_MULT_BAND_1','REFLECTANCE_MULT_BAND_3','REFLECTANCE_MULT_BAND_4']:
					if b[-1]=='1':
						ws.cell(row=k+54,column=2).value=text[let+26:let+36]
					if b[-1]=='3':
						ws.cell(row=k+54,column=3).value=text[let+26:let+36]
					if b[-1]=='4':
						ws.cell(row=k+54,column=4).value=text[let+26:let+36]
os.chdir('C:/Users/administrator/Desktop/Portland Greening/Bulk Order 903560/Landsat 4-5 TM C1 Level-1/Ims')						
wb.save('Portland Greening 2.xlsx')
