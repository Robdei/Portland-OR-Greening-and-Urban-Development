from __future__ import print_function
from statsmodels.compat import lzip
import statsmodels
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import statsmodels.stats.api as sms
import matplotlib as plt
import statsmodels.api as sm
from scipy.linalg import toeplitz
import glob, os
import openpyxl



files=[]
for i in glob.glob('*.tif'):
	files.append(i)
Direction=['E','NE','N','NW','W','SW','S','SE']
# Load data
wb=openpyxl.load_workbook('TempVDistance 2.xlsx')
ws=wb['Stats']
for i in range(len(files)):
	for d in range(len(Direction)):
		print(files[i][:-4]+Direction[d])
		sheet=files[i][:-4]+Direction[d]
		dat = pd.read_excel('TempVDistance 2.xlsx', sheetname=sheet)
		Dis=dat['Distance'].tolist()
		T=dat['Temp'].tolist()
# Fit regression model (using the natural log of one of the regressaors)
		results = smf.ols('Temp ~ Distance', data=dat).fit()

		name = ['Lagrange multiplier statistic', 'p-value', 'f-value', 'f p-value']

#BP Test for HeteroScadicity
		test = sms.het_breushpagan(results.resid, results.model.exog)
		p_value=lzip(name, test)[1][1]



#print(p_value)
		if p_value<.05:
			Dis = sm.add_constant(Dis)
			ols_resid = sm.OLS(T, Dis).fit().resid
			res_fit = sm.OLS(ols_resid[1:], ols_resid[:-1]).fit()
			rho = res_fit.params
			order = toeplitz(range(len(ols_resid)))
			sigma = rho**order
			gls_model = sm.GLS(T, Dis, sigma=sigma)
			gls_results = gls_model.fit()
			ws.cell(row=3*i+1, column=d+2).value=gls_results.params[1]
			ws.cell(row=3*i+2, column=d+2).value=gls_results.rsquared
			ws.cell(row=3*i+3, column=d+2).value=gls_results.pvalues[1]
		else:
			ws.cell(row=3*i+1, column=d+2).value=results.params[1]
			ws.cell(row=3*i+2, column=d+2).value=results.rsquared
			ws.cell(row=3*i+3, column=d+2).value=results.pvalues[1]
wb.save('Heat Island.xlsx')
